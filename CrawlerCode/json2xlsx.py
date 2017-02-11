# coding: utf-8

import os

import openpyxl

import config
import os_utils

try:
    import simplejson as json
except ImportError:
    import json


def write_to_excel(list_data, workbook, worksheet_name, number_of_sheets_written, date):
    """
    Write data to excel file
    :param list_data: list containing facebook page's info
    :param workbook: the workbook where to write the data
    :param worksheet_name: the name of the new worksheet
    :param number_of_sheets_written: The number of written sheets
    :param date: The date to print in the Excel file (launch date)
    :return:
    """
    # Create a new worksheet
    worksheet = workbook.create_sheet(title=worksheet_name.decode('utf-8'))
    # Start from the first cell. Rows and columns are one indexed.
    row = 2
    col = 2
    # The first column items : the header
    first_column = ['Page', 'Category', 'Likes', 'Talking about count', '% Talking about', 'Total posts',
                    'Max reactions/posts', 'Avg reactions/posts', 'Max comments/posts', 'Avg comments/posts',
                    'Max shares/posts', 'Avg shares/posts']
    # Write the first column
    for value in first_column:
        worksheet.cell(row=row, column=col, value=value)
        col += 1
    # Return to the second column
    col = 2
    # Variable we can use later
    max_avg_reactions = min_avg_reactions = list_data[0]['avg_reactions']
    max_avg_comments = min_avg_comments = list_data[0]['avg_comments']
    max_avg_shares = min_avg_shares = list_data[0]['avg_shares']
    max_talking_about_percent = min_talking_about_percent = list_data[0]['talking_about_percent']
    total_likes = 0
    total_talking_about_count = 0
    total_pages = 0
    # Write the data
    for data in list_data:
        if data is None:
            # Go to the next iteration
            continue
        # Jump to the next row and write all the data
        row += 1
        worksheet.cell(row=row, column=col, value=data['name'])
        worksheet.cell(row=row, column=col + 1, value=data['category'])
        worksheet.cell(row=row, column=col + 2, value=data['likes'])
        worksheet.cell(row=row, column=col + 3, value=data['talking_about_count'])
        worksheet.cell(row=row, column=col + 4, value=round(data['talking_about_percent'], 3))
        worksheet.cell(row=row, column=col + 5, value=data['total_posts'])
        worksheet.cell(row=row, column=col + 6, value=data['max_reactions'])
        worksheet.cell(row=row, column=col + 7, value=round(data['avg_reactions'], 3))
        worksheet.cell(row=row, column=col + 8, value=data['max_comments'])
        worksheet.cell(row=row, column=col + 9, value=round(data['avg_comments'], 3))
        worksheet.cell(row=row, column=col + 10, value=data['max_shares'])
        worksheet.cell(row=row, column=col + 11, value=round(data['avg_shares'], 3))
        # Increment the counters
        total_likes += data['likes']
        total_talking_about_count += data['talking_about_count']
        total_pages += 1
        # Calculate the intervals of the percentages of talking about
        if data['talking_about_percent'] < min_talking_about_percent:
            min_talking_about_percent = data["talking_about_percent"]
        elif data['talking_about_percent'] > max_talking_about_percent:
            max_talking_about_percent = data['talking_about_percent']
        # Calculate the intervals of the reactions
        if data['avg_reactions'] < min_avg_reactions:
            min_avg_reactions = data['avg_reactions']
        elif data['avg_reactions'] > max_avg_reactions:
            max_avg_reactions = data['avg_reactions']
        # Calculate the intervals of the percentages of the comments
        if data['avg_comments'] < min_avg_comments:
            min_avg_comments = data['avg_comments']
        elif data['avg_comments'] > max_avg_comments:
            max_avg_comments = data['avg_comments']
        # Calculate the intervals of the percentages of the shares
        if data['avg_shares'] < min_avg_shares:
            min_avg_shares = data['avg_shares']
        elif data['avg_shares'] > max_avg_shares:
            max_avg_shares = data['avg_shares']
    # The header for the second schedule
    first_column_all = ['Brand name', 'Total pages', 'Total likes', 'Total talking about count',
                        '% Talking about', 'Avg reactions/post', 'Avg comments/post', 'Avg shares/post']
    # Write the first column
    for value in first_column_all:
        worksheet.cell(row=row + 3, column=col, value=value)
        col += 1
    # Initialize the column and row number to start writing in the right cell
    col = 2
    row += 4
    # Write the data
    worksheet.cell(row=row, column=col, value=worksheet_name)
    worksheet.cell(row=row, column=col + 1, value=total_pages)
    worksheet.cell(row=row, column=col + 2, value=total_likes)
    worksheet.cell(row=row, column=col + 3, value=total_talking_about_count)
    # If there is only one page
    if len(list_data) == 1:
        worksheet.cell(row=row, column=col + 4, value=round(max_talking_about_percent, 3))
        worksheet.cell(row=row, column=col + 5, value=round(max_avg_reactions, 3))
        worksheet.cell(row=row, column=col + 6, value=round(max_avg_comments, 3))
        worksheet.cell(row=row, column=col + 7, value=round(max_avg_shares, 3))
    # Else
    else:
        worksheet.cell(row=row, column=col + 4,
                       value='[{} ; {}]'.format(round(min_talking_about_percent, 3),
                                                round(max_talking_about_percent, 3)))
        worksheet.cell(row=row, column=col + 5,
                       value='[{} ; {}]'.format(round(min_avg_reactions, 3), round(max_avg_reactions, 3)))
        worksheet.cell(row=row, column=col + 6,
                       value='[{} ; {}]'.format(round(min_avg_comments, 3), round(max_avg_comments, 3)))
        worksheet.cell(row=row, column=col + 7,
                       value='[{} ; {}]'.format(round(min_avg_shares, 3), round(max_avg_shares, 3)))
    # Jump to the Comparatif sheet
    worksheet = workbook.get_sheet_by_name('Comparatif')
    # Initialize the column and row values to write in the appropriate cell
    col = 2
    row = 3 + number_of_sheets_written
    # Write the data
    worksheet.cell(row=row, column=col, value=worksheet_name)
    worksheet.cell(row=row, column=col + 1, value=total_pages)
    worksheet.cell(row=row, column=col + 2, value=total_likes)
    worksheet.cell(row=row, column=col + 3, value=total_talking_about_count)
    # If there is only one page
    if len(list_data) == 1:
        worksheet.cell(row=row, column=col + 4, value=round(max_talking_about_percent, 3))
        worksheet.cell(row=row, column=col + 5, value=round(max_avg_reactions, 3))
        worksheet.cell(row=row, column=col + 6, value=round(max_avg_comments, 3))
        worksheet.cell(row=row, column=col + 7, value=round(max_avg_shares, 3))
    # Else
    else:
        worksheet.cell(row=row, column=col + 4,
                       value='[{} ; {}]'.format(round(min_talking_about_percent, 3),
                                                round(max_talking_about_percent, 3)))
        worksheet.cell(row=row, column=col + 5,
                       value='[{} ; {}]'.format(round(min_avg_reactions, 3), round(max_avg_reactions, 3)))
        worksheet.cell(row=row, column=col + 6,
                       value='[{} ; {}]'.format(round(min_avg_comments, 3), round(max_avg_comments, 3)))
        worksheet.cell(row=row, column=col + 7,
                       value='[{} ; {}]'.format(round(min_avg_shares, 3), round(max_avg_shares, 3)))
    # Writing the date
    worksheet.cell(row=row, column=col + 8,
                   value=date.strftime('%d/%m/%Y'))


def export_to_excel(path, filename, date):
    """
    Export stats to Excel
    :param path: The path of the data directory
    :param filename: The Excel filename
    :param date: The date to print (today's date)
    :return: Nothing
    """
    # Create a workbook
    workbook = openpyxl.Workbook()
    # Get the active worksheet
    worksheet = workbook.active
    worksheet.title = 'Comparatif'
    # The header of the file
    first_column_all = ['Brand name', 'Total pages', 'Total likes', 'Total talking about count',
                        '% Talking about', 'Avg reactions/post', 'Avg comments/post', 'Avg shares/post', 'Date']
    # Initialize to the cell at (1, 1) coordinates
    col = 2
    row = 2
    # Write the first column
    for value in first_column_all:
        worksheet.cell(row=row, column=col, value=value)
        col += 1
    # The total number of sheets written
    number_of_sheets_written = 0
    # Get the list of all directories in the current one
    dirs = os_utils.get_directory_list(path)
    # For each directory
    for directory in dirs:
        # Create a list where to store all the data
        storage = []
        # Get the list of all files in the directory
        subdirs = os_utils.get_directory_list(os.path.join(path, directory))
        # For each subdirectory
        for subdir in subdirs:
            files = os_utils.get_file_list(os.path.join(path, directory, subdir))
            # For each file in files
            for file in files:
                # If files doesn't contain the word "post" then continue
                if file.endswith('.json'):
                    # Extract data
                    data = os_utils.load_data_from_json_file(os.path.join(path, directory, subdir, file))
                    # Append the data to the list
                    storage.append(data)
        # Write the data to Excel file
        if len(storage) > 0:
            write_to_excel(storage, workbook, directory.split('_-_')[0], number_of_sheets_written, date)
        # Increment the number of written sheets
        number_of_sheets_written += 1
    # Save the workbook
    workbook.save(os.path.join(config.excel_dir, filename + '.xlsx'))
